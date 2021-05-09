from datetime import datetime
from openpyxl import Workbook
import requests
import json
import time
from openpyxl.utils import get_column_letter


def get_bed_details():
    district_response = requests.get('https://tncovidbeds.tnega.org/api/district')
    districts = json.loads(district_response.text)
    payload = {'searchString': '', 'sortCondition': {'Name': 1}, 'pageNumber': 1, 'pageLimit': 100000,
               'SortValue': 'Availability', 'BrowserId': 'b4c5b065a84c7d2b60e8b23d415b2c3a',
               'Districts': [], "IsGovernmentHospital": True, 'IsPrivateHospital': True,
               'FacilityTypes': ['CHO', 'CHC', 'CCC']}
    [payload['Districts'].append(district['id']) for district in districts['result']]
    headers = {'Content-Type': 'application/json', 'Accept': 'application/json'}
    hospital_response = requests.post('https://tncovidbeds.tnega.org/api/hospitals', json=payload, headers=headers)
    hospitals = json.loads(hospital_response.text)
    covid_centers = []
    for hospital in hospitals['result']:
        try:
            CovidBedDetails = hospital['CovidBedDetails']
            center = {'Name': hospital['Name'], 'District': hospital['District']['Name'],
                      'FacilityType': hospital['FacilityType'],
                      'Type': hospital['Type']['Name'], 'Landline': hospital['Landline'],
                      'MobileNumber': hospital['MobileNumber'],
                      'PrimaryContactPerson': hospital['PrimaryContactPerson'],
                      'TotalBeds': CovidBedDetails['TotalBedsInHospital'],
                      'BedsAllotedForCovidTreatment': CovidBedDetails['BedsAllotedForCovidTreatment'],
                      'AllotedO2Beds': CovidBedDetails['AllotedO2Beds'],
                      'AllotedNonO2Beds': CovidBedDetails['AllotedNonO2Beds'],
                      'AllotedICUBeds': CovidBedDetails['AllotedICUBeds'],
                      'OccupancyO2Beds': CovidBedDetails['OccupancyO2Beds'],
                      'OccupancyNonO2Beds': CovidBedDetails['OccupancyNonO2Beds'],
                      'OccupancyICUBeds': CovidBedDetails['OccupancyICUBeds'],
                      'VaccantO2Beds': CovidBedDetails['VaccantO2Beds'],
                      'VaccantNonO2Beds': CovidBedDetails['VaccantNonO2Beds'],
                      'VaccantICUBeds': CovidBedDetails['VaccantICUBeds'],
                      'StatusAsOf': CovidBedDetails['StatusAsOf'],
                      'TotalVaccantBeds': CovidBedDetails['TotalVaccantBeds'],
                      'UpdatedOn': datetime.fromtimestamp(CovidBedDetails['UpdatedOn'])
                      }
            contactdetails = ''
            for contact in hospital['ContactDetails']:
                if 'ContactNumber' in contact.keys():
                    contactdetails += contact['ContactNumber']
                if 'ContactName' in contact.keys():
                    if not (contact['ContactName'] is None or contact['ContactName'] == ''):
                        contactdetails += ' (' + contact['ContactName'] + ')'
                if 'Timing' in contact.keys():
                    contactdetails += ' [' + contact['Timing'] + ']'

                contactdetails += '\n'

                center['ContactDetails'] = contactdetails

            interested_keys = ['Line1', 'Line2', 'Line3', 'Line4', 'Taluk']
            address = ''
            for addresskey in hospital['AddressDetail'].keys():
                if addresskey in interested_keys:
                    if addresskey == 'Taluk':
                        address += hospital['AddressDetail']['Taluk']['Name'] + '\n'
                        continue
                    address += hospital['AddressDetail'][addresskey] + '\n'
            center['AddressDetails'] = address
            covid_centers.append(center)
        except:
            pass
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Hospital Name'
    sheet['B1'] = 'District'
    sheet['C1'] = 'Mobile Number'
    sheet['D1'] = 'Landline'
    sheet['E1'] = 'Primary Contact Person'
    sheet['F1'] = 'Contact Details'
    sheet['G1'] = 'Address Details'
    sheet['H1'] = 'FacilityType'
    sheet['I1'] = 'Type'
    sheet['J1'] = 'Total Beds'
    sheet['K1'] = 'Beds Alloted For CovidTreatment'
    sheet['L1'] = 'Alloted Oxygen Beds'
    sheet['M1'] = 'Occupancy Oxygen Beds'
    sheet['N1'] = 'Vaccant Oxygen Beds'
    sheet['O1'] = 'Alloted Non Oxygen Beds'
    sheet['P1'] = 'Occupancy Non Oxygen Beds'
    sheet['Q1'] = 'Vaccant Non Oxygen Beds'
    sheet['R1'] = 'Alloted ICU Beds'
    sheet['S1'] = 'Occupancy ICU Beds'
    sheet['T1'] = 'Vaccant ICU Beds'
    sheet['U1'] = 'Total Vaccant Beds'
    sheet['V1'] = 'Last Update'
    i = 2
    for center in covid_centers:
        sheet['A' + str(i)] = center['Name']
        sheet['B' + str(i)] = center['District']
        sheet['C' + str(i)] = center['MobileNumber']
        sheet['D' + str(i)] = center['Landline']
        sheet['E' + str(i)] = center['PrimaryContactPerson']
        sheet['F' + str(i)] = center['ContactDetails']
        sheet['G' + str(i)] = center['AddressDetails']
        sheet['H' + str(i)] = center['FacilityType']
        sheet['I' + str(i)] = center['FacilityType']
        sheet['J' + str(i)] = center['TotalBeds']
        sheet['K' + str(i)] = center['BedsAllotedForCovidTreatment']
        sheet['L' + str(i)] = center['AllotedO2Beds']
        sheet['M' + str(i)] = center['OccupancyO2Beds']
        sheet['N' + str(i)] = center['VaccantO2Beds']
        sheet['O' + str(i)] = center['AllotedNonO2Beds']
        sheet['P' + str(i)] = center['OccupancyNonO2Beds']
        sheet['Q' + str(i)] = center['VaccantNonO2Beds']
        sheet['R' + str(i)] = center['AllotedICUBeds']
        sheet['S' + str(i)] = center['OccupancyICUBeds']
        sheet['T' + str(i)] = center['VaccantICUBeds']
        sheet['U' + str(i)] = center['TotalVaccantBeds']
        sheet['V' + str(i)] = center['UpdatedOn']
        i = i + 1
    column_widths = []
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = column_width
    try:
        workbook.save(filename="tn-covid-bedtracker.xlsx")
        print('Updated the Excel on ' + str(datetime.now()))
    except:
        print('Failed to Update the Excel on '+ str(datetime.now()))

if __name__ == '__main__':


    nexttime = time.time()
    while True:
        get_bed_details()  # take t sec
        nexttime += 10
        sleeptime = nexttime - time.time()
        if sleeptime > 0:
            time.sleep(sleeptime)


